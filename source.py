import pandas as pd
import numpy as np
import datetime
from datetime import datetime, timedelta

# For Excel integration
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows

# For HTML dashboard
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# For HTML templating
from jinja2 import Template

# For logging
import logging
import os

# Jinja2 template for HTML report (moved outside function for clarity and efficiency if called multiple times)
REPORT_TEMPLATE = Template("""
<!DOCTYPE html>
<html>
<head>
    <title>{{ title }}</title>
    <style>
        body { font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; color: #333; }
        h1 { color: #008080; border-bottom: 2px solid #008080; padding-bottom: 10px; }
        h2 { color: #008080; margin-top: 30px; }
        .metadata { font-size: 12px; margin-bottom: 20px; color: #666; }
        .ai-flag { background: #FFF3CD; padding: 8px; border-radius: 4px; border-left: 4px solid #FFC107; margin: 10px 0; font-weight: bold; color: #8B4513;}
        table { border-collapse: collapse; width: 100%; margin: 20px 0; box-shadow: 0 2px 10px rgba(0,0,0,0.05); }
        th { background: #008080; color: white; padding: 12px 15px; text-align: left; border-bottom: 1px solid #ddd; }
        td { border: 1px solid #ddd; padding: 10px 15px; text-align: center; }
        tr:nth-child(even) { background: #f9f9f9; }
        .positive { color: green; font-weight: bold; }
        .negative { color: red; font-weight: bold; }
        .neutral { color: gray; }
    </style>
</head>
<body>
    <h1>{{ title }}</h1>
    <div class="metadata">
        Generated: {{ timestamp }} | Models: {{ models }} | Report ID: {{ report_id }}
    </div>
    <div class="ai-flag">
        This report contains AI-generated content. All data should be verified before use in client communications.
    </div>

    <h2>{{ section_title }}</h2>
    <table>
        <thead>
            <tr>
                {% for col in columns %}
                    <th>{{ col }}</th>
                {% endfor %}
            </tr>
        </thead>
        <tbody>
            {% for row in data %}
                <tr>
                    {% for val in row %}
                        <td
                          {% if loop.index0 == score_col_idx %}
                            class="{{ 'positive' if val > 0.3 else ('negative' if val < -0.3 else 'neutral') }}"
                          {% endif %}
                        >
                          {% if loop.index0 == score_col_idx %}
                            {{ val | round(3) }}
                          {% else %}
                            {{ val }}
                          {% endif %}
                        </td>
                    {% endfor %}
                </tr>
            {% endfor %}
        </tbody>
    </table>

    <div class="metadata">
        Compliance: CFA Standard V(B) - AI content flagged | Human review required before distribution.
    </div>
</body>
</html>
""")


def generate_synthetic_data(num_companies=5, num_obligors=4, num_days=10):
    """
    Generates synthetic datasets for sentiment, credit risk, and ESG governance.
    """
    np.random.seed(42)
    companies = ['AAPL', 'MSFT', 'GOOG', 'AMZN', 'TSLA'][:num_companies]
    obligors = ['Company A', 'Company B', 'Company C', 'Company D'][:num_obligors]
    today = datetime.now().date()

    # --- Sentiment Data ---
    sentiment_data = []
    for ticker in companies:
          for i in range(num_days):
            date = today - timedelta(days=num_days - 1 - i)
            sentiment_score = np.random.uniform(-0.8, 0.8)
            n_headlines = np.random.randint(5, 50)
            signal = 'Neutral'
            if sentiment_score > 0.3:
                  signal = 'Positive'
            elif sentiment_score < -0.3:
                  signal = 'Negative'
            sentiment_data.append({
                  'ticker': ticker,
                'date': date,
                'sentiment_score': sentiment_score,
                'n_headlines': n_headlines,
                'signal': signal
            })
    sentiment_df = pd.DataFrame(sentiment_data)

    # Ensure current day's sentiment is available for the dashboard
    current_sentiment_df = sentiment_df[sentiment_df['date'] == today].copy()
    if current_sentiment_df.empty:
          # If today's date not in generated, take the latest for each ticker
        current_sentiment_df = sentiment_df.loc[sentiment_df.groupby('ticker')['date'].idxmax()].copy()

    # --- Credit Data ---
    credit_data = []
    for obligor in obligors:
        pd_ensemble = np.random.uniform(0.01, 0.40) # Probability of Default
        rating_implied = ''
        if pd_ensemble < 0.03: rating_implied = 'AAA'
        elif pd_ensemble < 0.08: rating_implied = 'AA'
        elif pd_ensemble < 0.15: rating_implied = 'A'
        elif pd_ensemble < 0.25: rating_implied = 'BBB'
        elif pd_ensemble < 0.35: rating_implied = 'BB'
        else: rating_implied = 'B'

        model_agreement = f"{np.random.randint(1, 4)}/3" # e.g., 3/3 for high agreement

        credit_data.append({
              'obligor': obligor,
            'pd_ensemble': pd_ensemble,
            'rating_implied': rating_implied,
            'model_agreement': model_agreement
        })
    credit_df = pd.DataFrame(credit_data)

    # --- ESG Governance Data ---
    esg_governance_data = []
    for ticker in companies:
        ceo_comp = np.random.randint(5, 50) * 100000 # in USD
        board_size = np.random.randint(5, 15)
        esg_in_comp = np.random.choice(['Yes', 'No'], p=[0.7, 0.3]) # ESG metrics in CEO compensation
        esg_governance_data.append({
              'ticker': ticker,
            'ceo_total_compensation': ceo_comp,
            'board_size': board_size,
            'esg_in_compensation': esg_in_comp,
            'independent_directors': np.random.randint(3, board_size),
            'board_independence_pct': np.random.uniform(0.3, 0.9),
            'say_on_pay_approval': np.random.uniform(0.7, 0.99),
            'clawback_policy': np.random.choice(['Yes', 'No'], p=[0.8, 0.2])
        })
    esg_governance_df = pd.DataFrame(esg_governance_data)

    # For Plotly Radar Chart, we need aggregated ESG scores
    esg_scores = []
    for ticker in companies:
        esg_scores.append({
              'ticker': ticker,
            'environmental_score': np.random.uniform(30, 90),
            'social_score': np.random.uniform(30, 90),
            'governance_score': np.random.uniform(30, 90)
        })
    esg_overall_df = pd.DataFrame(esg_scores)

    return current_sentiment_df, sentiment_df, credit_df, esg_governance_df, esg_overall_df

def create_ai_dashboard_excel(sentiment_df_current, credit_df, output_path='ai_portfolio_dashboard.xlsx', output_dir='output_reports'):
    """
    Creates a formatted Excel workbook with AI model outputs for sentiment and credit risk,
    including conditional formatting, charts, and compliance metadata.
    """
    wb = openpyxl.Workbook()

    # --- Styles for professional formatting ---
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='008080', fill_type='solid') # Dark blue
    red_fill = PatternFill(start_color='FFC7CE', fill_type='solid') # Light Red
    green_fill = PatternFill(start_color='C6EFCE', fill_type='solid') # Light Green
    yellow_fill = PatternFill(start_color='FFEB9C', fill_type='solid') # Light Yellow (for neutral sentiment)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Sheet 1: Portfolio Sentiment ---
    ws1 = wb.active
    ws1.title = "Portfolio Sentiment"

    # Title and Metadata Header
    ws1['A1'] = 'Daily AI-Enhanced Portfolio Sentiment Dashboard'
    ws1['A1'].font = Font(bold=True, size=16, color='008080') # Dark blue title
    ws1['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Model: FinBERT | AI-Generated'
    ws1['A2'].font = Font(italic=True, size=10, color='666666')

    # Write data headers and apply styling
    # sentiment_df_current should only have one row per ticker for the "current" view.
    # We remove the 'date' column for this aggregated view.
    display_sentiment_df = sentiment_df_current[['ticker', 'sentiment_score', 'n_headlines', 'signal']].copy()

    rows = dataframe_to_rows(display_sentiment_df, index=False, header=True)
    for r_idx, row in enumerate(rows, 4): # Start writing data from row 4
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 4: # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    for col_idx, col_name in enumerate(display_sentiment_df.columns, 1):
          ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(str(col_name)), display_sentiment_df[col_name].astype(str).map(len).max()) + 2

    # Conditional Formatting for sentiment_score
    # Identify the column containing 'sentiment_score'
    score_col_letter = ''
    for col_idx, col_name in enumerate(display_sentiment_df.columns, 1):
          if col_name == 'sentiment_score':
            score_col_letter = openpyxl.utils.get_column_letter(col_idx)
            break

    last_row_ws1 = ws1.max_row
    if score_col_letter:
        # Green for Positive: score > 0.3
        ws1.conditional_formatting.add(
              f'{score_col_letter}5:{score_col_letter}{last_row_ws1}',
            CellIsRule(operator='greaterThan', formula=['0.3'], fill=green_fill)
        )
        # Red for Negative: score < -0.3
        ws1.conditional_formatting.add(
              f'{score_col_letter}5:{score_col_letter}{last_row_ws1}',
            CellIsRule(operator='lessThan', formula=['-0.3'], fill=red_fill)
        )
        # Yellow for Neutral: -0.3 <= score <= 0.3
        ws1.conditional_formatting.add(
              f'{score_col_letter}5:{score_col_letter}{last_row_ws1}',
            CellIsRule(operator='between', formula=['-0.3', '0.3'], fill=yellow_fill)
        )

    # Add sentiment bar chart
    chart1 = BarChart()
    chart1.title = "Sentiment by Company"
    chart1.y_axis.title = "Sentiment Score"
    chart1.x_axis.title = "Company Ticker"

    data_ref = Reference(ws1, min_col=display_sentiment_df.columns.get_loc('sentiment_score')+1,
                         min_row=5, max_row=last_row_ws1, max_col=display_sentiment_df.columns.get_loc('sentiment_score')+1)
    cats_ref = Reference(ws1, min_col=1, min_row=5, max_row=last_row_ws1)

    chart1.add_data(data_ref, titles_from_data=False)
    chart1.set_categories(cats_ref)
    chart1.width = 20
    chart1.height = 12

    ws1.add_chart(chart1, f"A{last_row_ws1 + 3}") # Place chart below data

    # --- Sheet 2: Credit Risk ---
    ws2 = wb.create_sheet("Credit Risk")

    # Title and Metadata Header
    ws2['A1'] = 'AI-Derived Credit Default Probabilities'
    ws2['A1'].font = Font(bold=True, size=16, color='008080')
    ws2['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Model: Stacking Ensemble | AI-Generated'
    ws2['A2'].font = Font(italic=True, size=10, color='666666')

    # Write data
    display_credit_df = credit_df[['obligor', 'pd_ensemble', 'rating_implied', 'model_agreement']].copy()

    rows = dataframe_to_rows(display_credit_df, index=False, header=True)
    for r_idx, row in enumerate(rows, 4):
          for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 4: # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                  cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    for col_idx, col_name in enumerate(display_credit_df.columns, 1):
          ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(str(col_name)), display_credit_df[col_name].astype(str).map(len).max()) + 2

    # Conditional Formatting for implied ratings
    # Define color mappings for ratings
    rating_colors = {
          'AAA': 'C6EFCE', # Light Green - Very Low Risk
        'AA': 'CCFFCC',
        'A': 'FFFFCC',  # Light Yellow - Low Risk
        'BBB': 'FFD966', # Amber - Medium Risk
        'BB': 'FFC000', # Orange - Elevated Risk
        'B': 'FF0000',  # Red - High Risk
        'CCC': '990000', # Dark Red - Very High Risk (though not in synthetic data, good to have)
        'CC': '660000',
        'C': '330000',
        'D': '000000'
    }

    rating_col_letter = ''
    for col_idx, col_name in enumerate(display_credit_df.columns, 1):
          if col_name == 'rating_implied':
            rating_col_letter = openpyxl.utils.get_column_letter(col_idx)
            break

    last_row_ws2 = ws2.max_row
    if rating_col_letter:
          for rating, hex_color in rating_colors.items():
            fill_color = PatternFill(start_color=hex_color, fill_type='solid')
            # Formula checks for exact match of rating string
            ws2.conditional_formatting.add(
                  f'{rating_col_letter}5:{rating_col_letter}{last_row_ws2}',
                CellIsRule(operator='equal', formula=[f'"{rating}"'], fill=fill_color)
            )

    # --- Sheet 3: Metadata & Compliance ---
    ws3 = wb.create_sheet("Metadata & Compliance")

    metadata = [
        ('Report Type', 'AI Model Output Dashboard'),
        ('Generation Date', datetime.now().isoformat()),
        ('Models Used', 'FinBERT (sentiment), Stacking Ensemble (credit)'),
        ('Data Sources', 'Financial News Feeds, LendingClub/Credit Data Providers'),
        ('AI-Generated Content', 'YES - All scores are model-generated'),
        ('Human Review Required', 'YES - Verify before client distribution'),
        ('Compliance Note', 'CFA Standard V(B): AI-assisted content flagged'),
        ('Disclaimer', 'This report contains AI-generated insights and should not be used as sole basis for investment decisions. Consult human experts.')
    ]

    for i, (key, val) in enumerate(metadata, 1):
        ws3.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws3.cell(row=i, column=2, value=val)
        ws3.column_dimensions[openpyxl.utils.get_column_letter(1)].width = max(len(str(m[0])) for m in metadata) + 2
        ws3.column_dimensions[openpyxl.utils.get_column_letter(2)].width = max(len(str(m[1])) for m in metadata) + 2

    # Save the workbook
    os.makedirs(output_dir, exist_ok=True) # Create directory if it doesn't exist
    full_output_path = os.path.join(output_dir, output_path)
    wb.save(full_output_path)
    logging.info(f"Excel dashboard saved: {full_output_path}")
    return full_output_path

def create_esg_governance_report(governance_df, output_path='esg_governance_report.xlsx', output_dir='output_reports'):
    """
    Creates a governance comparison report from ESG RAG extraction, with professional formatting.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Governance Comparison"

    # --- Styles ---
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='008080', fill_type='solid')
    green_fill_esg = PatternFill(start_color='C6EFCE', fill_type='solid') # Light Green for 'Yes'
    red_fill_esg = PatternFill(start_color='FFC7CE', fill_type='solid') # Light Red for 'No'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Professional header for the report
    ws.merge_cells('A1:H1') # Merge cells for the main title
    ws['A1'] = 'PORTFOLIO GOVERNANCE COMPARISON'
    ws['A1'].font = Font(bold=True, size=18, color='008080')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:H2') # Merge cells for source/date
    ws['A2'] = (f'Source: RAG extraction from DEF 14A proxy statements | '
                f'Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | AI-Assisted')
    ws['A2'].font = Font(italic=True, size=10, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Define columns to display and their order
    display_cols = [
        'ticker', 'ceo_total_compensation', 'board_size', 'independent_directors',
        'board_independence_pct', 'esg_in_compensation', 'say_on_pay_approval', 'clawback_policy'
    ]

    # Write data headers and apply styling
    for c_idx, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=4, column=c_idx, value=col_name.replace('_', ' ').title())
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Write data rows
    for r_idx, (_, row) in enumerate(governance_df.iterrows(), 5):
          for c_idx, col_name in enumerate(display_cols, 1):
            value = row.get(col_name, 'N/A')
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Highlight 'esg_in_compensation' column
            if col_name == 'esg_in_compensation':
                if str(value).lower() == 'yes':
                      cell.fill = green_fill_esg
                elif str(value).lower() == 'no':
                      cell.fill = red_fill_esg

    # Adjust column widths dynamically
    for col_idx, col_name in enumerate(display_cols, 1):
        max_length = max(len(str(row.get(col_name, 'N/A'))) for _, row in governance_df.iterrows())
        header_length = len(col_name.replace('_', ' ').title())
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(max_length, header_length) + 2

    # Add a portfolio summary row (example)
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value='PORTFOLIO SUMMARY').font = Font(bold=True, size=12)

    # Save the workbook
    os.makedirs(output_dir, exist_ok=True)
    full_output_path = os.path.join(output_dir, output_path)
    wb.save(full_output_path)
    logging.info(f"ESG report saved: {full_output_path}")
    return full_output_path

def create_interactive_dashboard(sentiment_df_current, sentiment_time_series_df, esg_overall_df, credit_df, output_path='ai_interactive_dashboard.html', output_dir='output_reports'):
    """
    Generates an interactive HTML dashboard combining multiple AI model outputs for daily monitoring.
    Panels: Sentiment bar chart, ESG radar chart, Sentiment Time Series, Credit Risk Heatmap.
    """
    # Create subplots: 2 rows, 2 columns
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=('Current News Sentiment by Company', 'ESG Scores by Pillar',
                        'Sentiment Trend Over Time', 'Credit Risk Distribution'),
        specs=[
              [{'type': 'bar'}, {'type': 'polar'}],
            [{'type': 'scatter'}, {'type': 'heatmap'}]
        ]
    )

    # Panel 1: Current Sentiment Bar Chart
    colors_sentiment = ['green' if s > 0.3 else 'red' if s < -0.3 else 'gray'
                        for s in sentiment_df_current['sentiment_score']]
    fig.add_trace(
        go.Bar(x=sentiment_df_current['ticker'],
               y=sentiment_df_current['sentiment_score'],
               marker_color=colors_sentiment,
               name='Current Sentiment'),
        row=1, col=1
    )

    # Panel 2: ESG Radar Chart (if esg_overall_df available)
    if esg_overall_df is not None and len(esg_overall_df) > 0:
        # Loop through a few companies for the radar chart
        for _, row in esg_overall_df.head(min(3, len(esg_overall_df))).iterrows():
              fig.add_trace(
                  go.Scatterpolar(
                      r=[row.get('environmental_score'), row.get('social_score'), row.get('governance_score'), row.get('environmental_score')], # Close the loop
                    theta=['Environmental', 'Social', 'Governance', 'Environmental'],
                    fill='toself',
                    opacity=0.5,
                    name=row.get('ticker', 'Unknown')
                ),
                row=1, col=2
            )

    # Panel 3: Sentiment Time Series Plot
    # Ensure sentiment_time_series_df is sorted by date
    sentiment_time_series_df_sorted = sentiment_time_series_df.sort_values(by=['ticker', 'date'])
    for ticker in sentiment_time_series_df_sorted['ticker'].unique():
        df_ticker = sentiment_time_series_df_sorted[sentiment_time_series_df_sorted['ticker'] == ticker]
        fig.add_trace(
              go.Scatter(x=df_ticker['date'], y=df_ticker['sentiment_score'],
                       mode='lines+markers', name=f'{ticker} Sentiment'),
            row=2, col=1
        )

    # Panel 4: Credit Risk Heatmap
    # For simplicity, let's make it a single row heatmap showing PDs.
    heatmap_data = credit_df[['obligor', 'pd_ensemble']].set_index('obligor').T

    fig.add_trace(
          go.Heatmap(
            z=heatmap_data.values,
            x=heatmap_data.columns,
            y=['Probability of Default'], # Single category for simplified view
            colorscale='Hot', # Example colorscale
            colorbar_title='PD Ensemble'
        ),
        row=2, col=2
    )

    # Update layout
    fig.update_layout(
        title_text=(f"AI Model Dashboard | Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | AI-Generated"),
        height=800,
        showlegend=True,
        template='plotly_white',
        hovermode='closest' # Better hover experience
    )

    # Update axes titles
    fig.update_xaxes(title_text="Company Ticker", row=1, col=1)
    fig.update_yaxes(title_text="Sentiment Score", row=1, col=1)
    fig.update_xaxes(title_text="Date", row=2, col=1)
    fig.update_yaxes(title_text="Sentiment Score", row=2, col=1)

    # Save the dashboard as an HTML file
    os.makedirs(output_dir, exist_ok=True)
    full_output_path = os.path.join(output_dir, output_path)
    fig.write_html(full_output_path, include_plotlyjs='cdn')
    logging.info(f"Interactive dashboard saved: {full_output_path}")
    return full_output_path

def generate_html_report_from_template(
    data_df,
    title,
    section_title,
    score_col_idx=1,
    output_path='sentiment_report.html',
    output_dir='output_reports'
):
    """
    Generates an HTML report from a DataFrame using a Jinja2 template, including compliance metadata.
    score_col_idx is 0-indexed for the DataFrame column that represents sentiment_score for conditional formatting.
    """
    html_content = REPORT_TEMPLATE.render(
        title=title,
        timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        models='FinBERT',
        report_id=f'RPT-{datetime.now().strftime("%Y%m%d%H%M%S")}',
        section_title=section_title,
        columns=data_df.columns.tolist(),
        data=data_df.values.tolist(),
        score_col_idx=score_col_idx
    )

    os.makedirs(output_dir, exist_ok=True)
    full_output_path = os.path.join(output_dir, output_path)

    with open(full_output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    logging.info(f"HTML report saved: {full_output_path}")
    return full_output_path


def generate_all_reports(
    num_companies=5,
    num_obligors=4,
    num_days=10,
    output_directory='output_reports',
    excel_dashboard_filename='ai_portfolio_dashboard.xlsx',
    esg_report_filename='esg_governance_report.xlsx',
    html_interactive_dashboard_filename='ai_interactive_dashboard.html',
    html_sentiment_report_filename='sentiment_report.html'
):
    """
    Orchestrates the generation of all AI-enhanced reports (Excel, interactive HTML, static HTML).

    Args:
        num_companies (int): Number of synthetic companies to generate.
        num_obligors (int): Number of synthetic obligors for credit data.
        num_days (int): Number of days for sentiment time series.
        output_directory (str): The directory where all reports will be saved.
        excel_dashboard_filename (str): Filename for the Excel AI dashboard.
        esg_report_filename (str): Filename for the ESG governance Excel report.
        html_interactive_dashboard_filename (str): Filename for the interactive Plotly HTML dashboard.
        html_sentiment_report_filename (str): Filename for the static HTML sentiment report.

    Returns:
        dict: A dictionary containing paths to all generated reports.
    """
    logging.info("Starting report generation process...")

    # 1. Generate Synthetic Data
    current_sentiment_df, sentiment_time_series_df, credit_df, esg_governance_df, esg_overall_df = \
        generate_synthetic_data(num_companies, num_obligors, num_days)
    logging.info("Synthetic data generated.")

    report_paths = {}

    # 2. Create AI Dashboard Excel
    excel_dashboard_path = create_ai_dashboard_excel(
        sentiment_df_current=current_sentiment_df,
        credit_df=credit_df,
        output_path=excel_dashboard_filename,
        output_dir=output_directory
    )
    report_paths['excel_ai_dashboard'] = excel_dashboard_path

    # 3. Create ESG Governance Excel Report
    esg_report_path = create_esg_governance_report(
        governance_df=esg_governance_df,
        output_path=esg_report_filename,
        output_dir=output_directory
    )
    report_paths['excel_esg_governance_report'] = esg_report_path

    # 4. Create Interactive Plotly HTML Dashboard
    html_dashboard_path = create_interactive_dashboard(
        sentiment_df_current=current_sentiment_df,
        sentiment_time_series_df=sentiment_time_series_df,
        esg_overall_df=esg_overall_df,
        credit_df=credit_df,
        output_path=html_interactive_dashboard_filename,
        output_dir=output_directory
    )
    report_paths['html_interactive_dashboard'] = html_dashboard_path

    # 5. Generate Static HTML Sentiment Report from Template
    html_report_path = generate_html_report_from_template(
        data_df=current_sentiment_df[['ticker', 'sentiment_score', 'n_headlines', 'signal']],
        title='AI-Enhanced Portfolio Sentiment Summary',
        section_title='Current Portfolio Sentiment Overview',
        score_col_idx=1, # 'sentiment_score' is the second column (index 1)
        output_path=html_sentiment_report_filename,
        output_dir=output_directory
    )
    report_paths['html_sentiment_summary'] = html_report_path

    logging.info("All reports generated successfully.")
    return report_paths

if __name__ == "__main__":
    # Configure basic logging if run directly
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    logging.info("Running report generation directly...")

    # Example of how to call the main function
    generated_report_paths = generate_all_reports(
        num_companies=5,
        num_obligors=4,
        num_days=10,
        output_directory='generated_reports_output' # Reports will be saved here
    )

    logging.info("\nGenerated Report Paths:")
    for report_type, path in generated_report_paths.items():
        logging.info(f"- {report_type}: {os.path.abspath(path)}")
